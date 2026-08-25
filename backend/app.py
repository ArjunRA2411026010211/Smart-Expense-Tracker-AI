from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    send_file,
    flash
)
import time

import random
import smtplib
import os
import mysql.connector
import joblib
from pathlib import Path


from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

from datetime import datetime
from functools import wraps
from io import BytesIO

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch


app = Flask(
    __name__,
    template_folder="../frontend/templates",
    static_folder="../frontend/static"
)

app.secret_key = os.environ.get(
    "SECRET_KEY",
    "dev-secret-key"
)


# ==================================================
# MYSQL CONNECTION
# ==================================================

db = mysql.connector.connect(
    host=os.environ.get(
        "MYSQLHOST",
        "localhost"
    ),
    user=os.environ.get(
        "MYSQLUSER",
        "root"
    ),
    password=os.environ.get(
        "MYSQLPASSWORD",
        ""
    ),
    database=os.environ.get(
        "MYSQLDATABASE",
        "smart_expense_tracker"
    ),
    port=int(
        os.environ.get(
            "MYSQLPORT",
            3306
        )
    )
)

# ==================================================
# LOGIN PROTECTION
# ==================================================

def login_required(function):

    @wraps(function)
    def decorated_function(*args, **kwargs):

        if "user_id" not in session:
            return redirect(url_for("login"))

        return function(*args, **kwargs)

    return decorated_function
def send_otp_email(receiver_email, otp):

    sender_email = os.environ.get("APP_EMAIL")
    app_password = os.environ.get("APP_EMAIL_PASSWORD")

    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = "Smart Expense Tracker - Password Reset OTP"

    body = f"""
Hello,

Your OTP for resetting your Smart Expense Tracker password is:

{otp}

Do not share this OTP with anyone.

Smart Expense Tracker AI
"""

    message.attach(MIMEText(body, "plain"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender_email, app_password)

        server.sendmail(
            sender_email,
            receiver_email,
            message.as_string()
        )


# ==================================================
# LOGIN PAGE
# ==================================================

@app.route("/")
def login():

    if "user_id" in session:
        return redirect(url_for("home"))

    return render_template("login.html")


# ==================================================
# CHECK LOGIN
# ==================================================

@app.route("/login", methods=["POST"])
def check_login():

    email = request.form["email"]
    password = request.form["password"]

    cursor = db.cursor(dictionary=True)

    cursor.execute("""
        SELECT id, fullname, email, phone, password
        FROM users
        WHERE email = %s
    """, (email,))

    user = cursor.fetchone()

    cursor.close()

    if user and check_password_hash(user["password"], password):

        session["user_id"] = user["id"]
        session["fullname"] = user["fullname"]
        session["email"] = user["email"]
        session["phone"] = user["phone"]

        return redirect(url_for("home"))

    return render_template(
        "login.html",
        error="Invalid Email or Password"
    )
# ==================================================
# OTP PAGE
# ==================================================

@app.route("/otp", methods=["GET", "POST"])
def otp():

    if "reset_email" not in session or "reset_otp" not in session:
        return redirect(url_for("forgot"))

    if request.method == "POST":

        entered_otp = request.form["otp"].strip()

        saved_otp = session.get("reset_otp")
        otp_created_at = session.get("otp_created_at")

        if not otp_created_at:
            return redirect(url_for("forgot"))

        # OTP expires after 5 minutes
        if time.time() - otp_created_at > 300:

            session.pop("reset_otp", None)
            session.pop("otp_created_at", None)

            return render_template(
                "otp.html",
                error="OTP expired. Please request a new OTP."
            )

        if entered_otp != saved_otp:

            return render_template(
                "otp.html",
                error="Invalid OTP. Please try again."
            )

        # OTP verified
        session["otp_verified"] = True

        return redirect(url_for("reset_password"))

    return render_template("otp.html")

# ==================================================
# RESEND OTP
# ==================================================

@app.route("/resend-otp")
def resend_otp():

    email = session.get("reset_email")

    if not email:
        return redirect(url_for("forgot"))

    # Generate new 6-digit OTP
    new_otp = str(random.randint(100000, 999999))

    # Replace old OTP and restart expiry timer
    session["reset_otp"] = new_otp
    session["otp_created_at"] = time.time()
    session.pop("otp_verified", None)

    try:

        send_otp_email(
            email,
            new_otp
        )

    except Exception as error:

        print("Resend OTP Error:", error)

        return render_template(
            "otp.html",
            error="Unable to resend OTP. Please try again."
        )

    return render_template(
        "otp.html",
        success="A new OTP has been sent to your email."
    )
# ==================================================
# RESET PASSWORD
# ==================================================

@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():

    if "reset_email" not in session:
        return redirect(url_for("forgot"))

    if not session.get("otp_verified"):
        return redirect(url_for("otp"))

    if request.method == "POST":

        new_password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        if new_password != confirm_password:

            return render_template(
                "reset-password.html",
                error="Passwords do not match."
            )

        hashed_password = generate_password_hash(new_password)

        cursor = db.cursor()

        cursor.execute("""
            UPDATE users
            SET password = %s
            WHERE email = %s
        """, (
            hashed_password,
            session["reset_email"]
        ))

        db.commit()
        cursor.close()

        session.pop("reset_email", None)
        session.pop("reset_otp", None)
        session.pop("otp_verified", None)

        flash(
            "Password reset successfully. Please login.",
            "success"
        )

        return redirect(url_for("login"))

    return render_template("reset-password.html")
# ==================================================
# FORGOT PASSWORD
# ==================================================
@app.route("/forgot-password", methods=["GET", "POST"])
def forgot():

    if request.method == "POST":

        email = request.form["email"].strip()

        cursor = db.cursor(dictionary=True)

        cursor.execute("""
            SELECT id, email
            FROM users
            WHERE email = %s
        """, (email,))

        user = cursor.fetchone()

        cursor.close()

        if not user:

            return render_template(
                "forgot-password.html",
                error="Email address not found."
            )

        # Generate 6-digit OTP
        otp_code = str(random.randint(100000, 999999))

        # Store email and OTP temporarily
        session["reset_email"] = email
        session["reset_otp"] = otp_code
        session["otp_created_at"] = time.time()

        try:

            send_otp_email(
                email,
                otp_code
            )

        except Exception as error:

            print("OTP Email Error:", error)

            return render_template(
                "forgot-password.html",
                error="Unable to send OTP. Please try again."
            )

        return redirect(url_for("otp"))

    return render_template("forgot-password.html")
# ==================================================
# HOME PAGE
# ==================================================

@app.route("/home")
@login_required
def home():

    user_id = session["user_id"]

    cursor = db.cursor(dictionary=True)

    # Total Income
    cursor.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM income
        WHERE user_id = %s
    """, (user_id,))

    total_income = float(cursor.fetchone()["total"])

    # Total Expenses
    cursor.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM expenses
        WHERE user_id = %s
    """, (user_id,))

    total_expense = float(cursor.fetchone()["total"])

    # Expense transaction count
    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM expenses
        WHERE user_id = %s
    """, (user_id,))

    expense_count = cursor.fetchone()["total"]

    # Income transaction count
    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM income
        WHERE user_id = %s
    """, (user_id,))

    income_count = cursor.fetchone()["total"]

    total_transactions = expense_count + income_count

    # Recent Expenses for Home Preview
    cursor.execute("""
        SELECT
            category,
            amount,
            description,
            expense_date
        FROM expenses
        WHERE user_id = %s
        ORDER BY id DESC
        LIMIT 4
    """, (user_id,))

    recent_expenses = cursor.fetchall()

    # Format Dates
    for expense in recent_expenses:
        if expense["expense_date"]:
            expense["expense_date"] = (
                expense["expense_date"].strftime("%d %b %Y")
            )

    cursor.close()

    savings = total_income - total_expense

    return render_template(
        "index.html",
        income=total_income,
        expenses=total_expense,
        savings=savings,
        total_transactions=total_transactions,
        recent_expenses=recent_expenses,
        fullname=session.get("fullname", "User")
    )


# ==================================================
# DASHBOARD
# ==================================================

@app.route("/dashboard")
@login_required
def dashboard():

    user_id = session["user_id"]

    current_month = datetime.now().month
    current_year = datetime.now().year

    cursor = db.cursor(dictionary=True)

    # ==================================================
    # TOTAL INCOME
    # ==================================================

    cursor.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM income
        WHERE user_id = %s
    """, (user_id,))

    total_income = float(
        cursor.fetchone()["total"]
    )


    # ==================================================
    # TOTAL EXPENSES
    # ==================================================

    cursor.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM expenses
        WHERE user_id = %s
    """, (user_id,))

    total_expense = float(
        cursor.fetchone()["total"]
    )


    # ==================================================
    # EXPENSE COUNT
    # ==================================================

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM expenses
        WHERE user_id = %s
    """, (user_id,))

    expense_count = (
        cursor.fetchone()["total"]
    )


    # ==================================================
    # INCOME COUNT
    # ==================================================

    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM income
        WHERE user_id = %s
    """, (user_id,))

    income_count = (
        cursor.fetchone()["total"]
    )

    total_transactions = (
        expense_count + income_count
    )


    # ==================================================
    # RECENT EXPENSES
    # ==================================================

    cursor.execute("""
        SELECT
            id,
            category,
            amount,
            description,
            expense_date
        FROM expenses
        WHERE user_id = %s
        ORDER BY id DESC
        LIMIT 5
    """, (user_id,))

    recent_expenses = (
        cursor.fetchall()
    )


    # ==================================================
    # RECENT INCOME
    # ==================================================

    cursor.execute("""
        SELECT
            id,
            source,
            amount,
            income_date
        FROM income
        WHERE user_id = %s
        ORDER BY id DESC
        LIMIT 5
    """, (user_id,))

    recent_income = (
        cursor.fetchall()
    )


    # ==================================================
    # FORMAT DATES
    # ==================================================

    for expense in recent_expenses:

        if expense["expense_date"]:

            expense["expense_date"] = (
                expense["expense_date"]
                .strftime("%d %b %Y")
            )


    for income_item in recent_income:

        if income_item["income_date"]:

            income_item["income_date"] = (
                income_item["income_date"]
                .strftime("%d %b %Y")
            )


    # ==================================================
    # CURRENT MONTH CATEGORY TOTALS
    # ==================================================

    cursor.execute("""
        SELECT
            category,
            SUM(amount) AS total,
            COUNT(*) AS transaction_count
        FROM expenses
        WHERE user_id = %s
          AND MONTH(expense_date) = %s
          AND YEAR(expense_date) = %s
        GROUP BY category
        ORDER BY total DESC
    """, (
        user_id,
        current_month,
        current_year
    ))

    monthly_category_rows = (
        cursor.fetchall()
    )


    monthly_category_summary = []

    for row in monthly_category_rows:

        monthly_category_summary.append({
            "category":
                row["category"].title(),

            "amount":
                float(row["total"]),

            "transaction_count":
                row["transaction_count"]
        })


    # ==================================================
    # MONTHLY TOTAL EXPENSE
    # ==================================================

    cursor.execute("""
        SELECT
            COALESCE(SUM(amount), 0) AS total
        FROM expenses
        WHERE user_id = %s
          AND MONTH(expense_date) = %s
          AND YEAR(expense_date) = %s
    """, (
        user_id,
        current_month,
        current_year
    ))

    monthly_total_expense = float(
        cursor.fetchone()["total"]
    )


    # ==================================================
    # CURRENT MONTH CHART
    # ==================================================

    chart_labels = []
    chart_values = []

    for row in monthly_category_rows:

        chart_labels.append(
            row["category"].title()
        )

        chart_values.append(
            float(row["total"])
        )


    # ==================================================
    # HIGHEST SPENDING CATEGORY THIS MONTH
    # ==================================================

    if monthly_category_rows:

        top_category = (
            monthly_category_rows[0]
        )

        top_amount = float(
            top_category["total"]
        )

        possible_saving = (
            top_amount * 0.15
        )

        ai_suggestion = (
            f"Your highest spending category this month is "
            f"{top_category['category'].title()} "
            f"at ₹{top_amount:,.2f}. "
            f"Reducing it by 15% could save approximately "
            f"₹{possible_saving:,.2f}."
        )

    else:

        ai_suggestion = (
            "No expenses recorded for this month yet."
        )


    cursor.close()


    # ==================================================
    # SAVINGS
    # ==================================================

    savings = (
        total_income - total_expense
    )

    current_month_name = datetime(
        current_year,
        current_month,
        1
    ).strftime("%B")


    # ==================================================
    # RENDER DASHBOARD
    # ==================================================

    return render_template(
        "dashboard.html",

        income=total_income,
        expenses=total_expense,
        savings=savings,

        total_transactions=total_transactions,

        recent_expenses=recent_expenses,
        recent_income=recent_income,

        ai_suggestion=ai_suggestion,

        chart_labels=chart_labels,
        chart_values=chart_values,

        monthly_category_summary=monthly_category_summary,
        monthly_total_expense=monthly_total_expense,

        current_month_name=current_month_name,
        current_year=current_year,

        fullname=session.get(
            "fullname",
            "User"
        )
    )

# ==================================================
# BANK STATEMENT UPLOAD PAGE
# ==================================================

@app.route("/upload-statement")
@login_required
def upload_statement_page():
    return render_template("upload-statement.html")
# ==================================================
# LOAD ML TRANSACTION MODEL
# ==================================================

BASE_DIR = Path(__file__).resolve().parent

MODEL_PATH = BASE_DIR / "ml" / "transaction_model.pkl"
VECTORIZER_PATH = BASE_DIR / "ml" / "vectorizer.pkl"

transaction_model = joblib.load(MODEL_PATH)
transaction_vectorizer = joblib.load(VECTORIZER_PATH)
# ==================================================
# MERCHANT CATEGORY MEMORY
# ==================================================

def get_saved_merchant_category(user_id, description):

    description = str(description or "").lower().strip()

    if not description:
        return None

    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT merchant_pattern, category
            FROM merchant_category_rules
            WHERE user_id = %s
            ORDER BY LENGTH(merchant_pattern) DESC
        """, (user_id,))

        rules = cursor.fetchall()

        for rule in rules:

            pattern = str(
                rule["merchant_pattern"]
            ).lower().strip()

            if pattern and pattern in description:

                print(
                    "Learned category:",
                    description,
                    "=>",
                    rule["category"]
                )

                return rule["category"]

    finally:
        cursor.close()

    return None


# ==================================================
# SAVE / UPDATE MERCHANT CATEGORY
# ==================================================

def save_merchant_category(user_id, description, category):

    description = str(description or "").strip()
    category = str(category or "").strip()

    if not description or not category:
        return

    # Normalize description so matching is consistent
    merchant_pattern = description.lower()

    cursor = db.cursor()

    try:

        cursor.execute("""
            INSERT INTO merchant_category_rules
                (user_id, merchant_pattern, category)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE
                category = VALUES(category)
        """, (
            user_id,
            merchant_pattern,
            category
        ))

        db.commit()

        print(
            "Merchant learned:",
            merchant_pattern,
            "=>",
            category
        )

    except Exception as error:

        db.rollback()

        print(
            "Merchant learning error:",
            error
        )

    finally:
        cursor.close()


# ==================================================
# TRANSACTION CATEGORIZATION
# ==================================================

def categorize_transaction(description, transaction_type):

    description = str(description or "").strip()
    transaction_type = str(transaction_type or "").lower().strip()

    text_lower = description.lower()

    # ==================================================
    # SALARY
    # ==================================================

    if any(word in text_lower for word in [
        "salary",
        "payroll",
        "wages",
        "stipend"
    ]):
        return "Salary"


    # ==================================================
    # FOOD
    # ==================================================

    if any(word in text_lower for word in [
        "swiggy",
        "zomato",
        "kfc",
        "mcdonald",
        "restaurant",
        "dominos",
        "pizza",
        "hotel",
        "cafe",
        "bakery"
    ]):
        return "Food"


    # ==================================================
    # PETROL
    # ==================================================

    if any(word in text_lower for word in [
        "indian oil",
        "iocl",
        "hpcl",
        "bharat petroleum",
        "bpcl",
        "petrol",
        "fuel",
        "diesel",
        "shell"
    ]):
        return "Petrol"


    # ==================================================
    # GROCERIES
    # ==================================================

    if any(word in text_lower for word in [
        "dmart",
        "d mart",
        "bigbasket",
        "blinkit",
        "zepto",
        "supermarket",
        "super store",
        "super stores",
        "grocery",
        "groceries",
        "provision",
        "mithran"
    ]):
        return "Groceries"


    # ==================================================
    # SHOPPING
    # ==================================================

    if any(word in text_lower for word in [
        "amazon",
        "flipkart",
        "myntra",
        "ajio",
        "meesho",
        "shopping",
        "clothing",
        "retail concepts",
        "retail",
        "electronics"
    ]):
        return "Shopping"


    # ==================================================
    # STUDIES
    # ==================================================

    if any(word in text_lower for word in [
        "college",
        "university",
        "school",
        "tuition",
        "course",
        "education",
        "institute",
        "exam fee",
        "academic"
    ]):
        return "Studies"


    # ==================================================
    # RENT
    # ==================================================

    if any(word in text_lower for word in [
        "rent",
        "landlord",
        "hostel",
        "pg rent",
        "house rent",
        "room rent"
    ]):
        return "Rent"


    # ==================================================
    # TRANSPORT
    # ==================================================

    if any(word in text_lower for word in [
        "uber",
        "ola",
        "rapido",
        "irctc",
        "railway",
        "metro",
        "taxi",
        "bus",
        "redbus"
    ]):
        return "Transport"


    # ==================================================
    # BILLS
    # ==================================================

    if any(word in text_lower for word in [
        "airtel",
        "jio",
        "recharge",
        "electricity",
        "broadband",
        "water bill",
        "gas bill",
        "utility",
        "gpayrecharge",
        "gpay.bp.utility"
    ]):
        return "Bills"


    # ==================================================
    # MEDICAL
    # ==================================================

    if any(word in text_lower for word in [
        "apollo",
        "pharmacy",
        "hospital",
        "medical",
        "doctor",
        "clinic",
        "medicine",
        "medplus",
        "diagnostic"
    ]):
        return "Medical"


    # ==================================================
    # ENTERTAINMENT
    # ==================================================

    if any(word in text_lower for word in [
        "netflix",
        "spotify",
        "hotstar",
        "bookmyshow",
        "cinema",
        "movie",
        "theatre",
        "prime video",
        "gaming"
    ]):
        return "Entertainment"


    # ==================================================
    # EMI
    # ==================================================

    if any(word in text_lower for word in [
        "emi",
        "loan installment",
        "loan payment",
        "finance installment"
    ]):
        return "EMI"


    # ==================================================
    # INVESTMENT
    # ==================================================

    if any(word in text_lower for word in [
        "zerodha",
        "groww",
        "mutual fund",
        "sip",
        "investment",
        "upstox",
        "fixed deposit"
    ]):
        return "Investment"

    # ==================================================
    # DETECT BANK / UPI TRANSFERS
    # ==================================================

    upi_indicators = [
        "upi",
        "ptmupi",
        "yblupi",
        "merupi",
        "mchupi",
        "@okicici",
        "@okaxis",
        "@oksbi",
        "@ybl",
        "@paytm",
        "@upi"
    ]

    is_upi_transfer = any(
        word in text_lower
        for word in upi_indicators
    )


    # ==================================================
    # CREDIT TRANSACTIONS
    # ==================================================

    if transaction_type in [
        "credit",
        "cr",
        "income"
    ]:

        # Salary was already detected above.

        # Real bank UPI credits from people should not
        # automatically become normal Income.
        if is_upi_transfer:
            return "Personal Transfer"

        # Bank descriptions often contain a person's
        # name after "/".
        if "/" in description:
            return "Personal Transfer"

        return "Income"


    # ==================================================
    # DEBIT PERSONAL TRANSFERS
    # ==================================================

    if transaction_type in [
        "debit",
        "dr",
        "expense"
    ]:

        # Known merchants were already detected by
        # Food/Groceries/Shopping/etc. rules above.

        if is_upi_transfer:
            return "Personal Transfer"

        # Real bank transfer narrations commonly look:
        # BANKCODE/NAME
        if "/" in description:
            return "Personal Transfer"


    # ==================================================
    # ML PREDICTION
    # ==================================================

    try:

        vector = transaction_vectorizer.transform(
            [description]
        )

        prediction = transaction_model.predict(
            vector
        )[0]

        probabilities = transaction_model.predict_proba(
            vector
        )[0]

        confidence = max(probabilities)

        print(
            "ML Prediction:",
            description,
            "=>",
            prediction,
            f"({confidence * 100:.1f}%)"
        )

        if confidence >= 0.70:
            return prediction

    except Exception as error:

        print(
            "ML Categorization Error:",
            error
        )


    # ==================================================
    # FINAL FALLBACK
    # ==================================================

    return "Other"
# ==================================================
# PREVIEW BANK STATEMENT CSV
# ==================================================
@app.route("/preview-statement", methods=["POST"])
@login_required
def preview_statement():

    statement_file = request.files.get("statement_file")

    # ==================================================
    # CHECK FILE
    # ==================================================

    if not statement_file or statement_file.filename == "":

        flash(
            "Please select a CSV, Excel, or PDF file.",
            "error"
        )

        return redirect(
            url_for("upload_statement_page")
        )

    filename = statement_file.filename.lower()

    if not filename.endswith((".csv", ".xlsx", ".pdf")):

        flash(
            "Only CSV, Excel (.xlsx), and PDF files are supported.",
            "error"
        )

        return redirect(
            url_for("upload_statement_page")
        )

    try:

        import csv
        import io
        import re
        from datetime import datetime

        rows = []
        original_columns = []

        # ==================================================
        # CSV
        # ==================================================

        if filename.endswith(".csv"):

            file_content = statement_file.stream.read().decode(
                "utf-8-sig"
            )

            csv_reader = csv.DictReader(
                io.StringIO(file_content)
            )

            rows = list(csv_reader)

            original_columns = csv_reader.fieldnames or []

        # ==================================================
        # EXCEL
        # ==================================================

        elif filename.endswith(".xlsx"):

            from openpyxl import load_workbook

            excel_bytes = statement_file.read()

            workbook = load_workbook(
                BytesIO(excel_bytes),
                read_only=True,
                data_only=True
            )

            worksheet = workbook.active

            excel_rows = list(
                worksheet.iter_rows(
                    values_only=True
                )
            )

            if excel_rows:

                original_columns = [
                    str(value).strip()
                    if value is not None
                    else ""
                    for value in excel_rows[0]
                ]

                for excel_row in excel_rows[1:]:

                    row_dictionary = {}

                    for index, column_name in enumerate(
                        original_columns
                    ):

                        if not column_name:
                            continue

                        value = (
                            excel_row[index]
                            if index < len(excel_row)
                            else ""
                        )

                        row_dictionary[column_name] = (
                            ""
                            if value is None
                            else str(value)
                        )

                    rows.append(
                        row_dictionary
                    )

        # ==================================================
        # PDF
        # ==================================================

        elif filename.endswith(".pdf"):

            import pdfplumber
            import re

            pdf_bytes = statement_file.read()

            rows = []

            original_columns = [
                "Date",
                "Description",
                "Amount",
                "Type"
            ]

            # Matches:
            # 06 May 2026
            # 07 May 2026
            date_pattern = re.compile(
                r"\b(\d{2}\s+[A-Za-z]{3}\s+\d{4})\b"
            )

            # Matches:
            # INR 8,000.00
            # INR 181.16
            amount_pattern = re.compile(
                r"INR\s+([\d,]+\.\d{2})",
                re.IGNORECASE
            )

            transaction_blocks = []

            # ==========================================
            # EXTRACT TEXT FROM ALL PDF PAGES
            # ==========================================

            with pdfplumber.open(
                BytesIO(pdf_bytes)
            ) as pdf:

                current_block = ""

                for page in pdf.pages:

                    text = page.extract_text()

                    if not text:
                        continue

                    for line in text.split("\n"):

                        line = line.strip()

                        if not line:
                            continue

                        # Skip headers / summary lines

                        if line.lower().startswith(
                            "date transaction details"
                        ):
                            continue

                        if line.lower().startswith(
                            "account statement"
                        ):
                            continue

                        if line.lower().startswith(
                            "ending balance"
                        ):
                            continue

                        if line.lower().startswith(
                            "total inr"
                        ):
                            continue

                        # ==================================
                        # NEW TRANSACTION STARTS WITH DATE
                        # ==================================

                        if date_pattern.search(line):

                            if current_block:

                                transaction_blocks.append(
                                    current_block.strip()
                                )

                            current_block = line

                        else:

                            # Multiline transaction description

                            if current_block:

                                current_block += " " + line

                # Save final transaction

                if current_block:

                    transaction_blocks.append(
                        current_block.strip()
                    )

            # ==========================================
            # PARSE EACH TRANSACTION BLOCK
            # ==========================================

            for block in transaction_blocks:

                date_match = date_pattern.search(
                    block
                )

                if not date_match:
                    continue

                transaction_date_text = (
                    date_match.group(1)
                )

                # Convert:
                # 06 May 2026
                # ->
                # 2026-05-06

                try:

                    transaction_date = (
                        datetime.strptime(
                            transaction_date_text,
                            "%d %b %Y"
                        ).strftime(
                            "%Y-%m-%d"
                        )
                    )

                except ValueError:

                    continue

                # ==========================================
                # FIND ALL INR VALUES
                # ==========================================

                amount_matches = (
                    amount_pattern.findall(
                        block
                    )
                )

                if len(amount_matches) < 2:
                    continue

                # In Indian Bank statement:
                #
                # Debit transaction:
                # INR 8,000.00 - INR 2,944.45
                #
                # Credit transaction:
                # - INR 1,500.00 INR 6,444.45
                #
                # Last INR value = balance

                balance_amount = (
                    amount_matches[-1]
                )

                transaction_amount = (
                    amount_matches[-2]
                )

                # ==========================================
                # DETECT DEBIT / CREDIT
                # ==========================================

                transaction_type = ""

                # Remove description/date first so we can
                # inspect the amount section more reliably.

                first_amount_position = None

                amount_search = re.search(
                    r"INR\s+[\d,]+\.\d{2}",
                    block,
                    re.IGNORECASE
                )

                if amount_search:

                    first_amount_position = (
                        amount_search.start()
                    )

                if first_amount_position is None:
                    continue

                amount_section = block[
                    first_amount_position:
                ]

                # Indian Bank layout:
                #
                # DEBIT:
                # INR amount - INR balance
                #
                # CREDIT:
                # - INR amount INR balance

                if re.match(
                    r"INR\s+[\d,]+\.\d{2}\s+-",
                    amount_section,
                    re.IGNORECASE
                ):

                    transaction_type = "Debit"

                elif re.match(
                    r"-\s*INR\s+[\d,]+\.\d{2}",
                    amount_section,
                    re.IGNORECASE
                ):

                    transaction_type = "Credit"

                else:

                    # Fallback detection

                    if " - INR " in amount_section:

                        transaction_type = "Debit"

                    else:

                        transaction_type = "Credit"

                # ==========================================
                # CLEAN DESCRIPTION
                # ==========================================

                description = block

                # Remove date

                description = description.replace(
                    transaction_date_text,
                    "",
                    1
                )

                # Remove amount section completely

                description = re.split(
                    r"\s+(?:-\s*)?INR\s+[\d,]+\.\d{2}",
                    description,
                    maxsplit=1,
                    flags=re.IGNORECASE
                )[0]

                # Clean whitespace

                description = re.sub(
                    r"\s+",
                    " ",
                    description
                ).strip(" |-")

                # ==========================================
                # ADD TRANSACTION
                # ==========================================

                if (
                    transaction_date
                    and description
                    and transaction_amount
                    and transaction_type
                ):

                    rows.append({

                        "Date":
                            transaction_date,

                        "Description":
                            description,

                        "Amount":
                            transaction_amount,

                        "Type":
                            transaction_type
                    })

            print(
                "PDF transactions extracted:",
                len(rows)
            )

        # ==================================================
        # CHECK ROWS
        # ==================================================

        if not rows:

            flash(
                "No transactions found in the statement.",
                "error"
            )

            return redirect(
                url_for("upload_statement_page")
            )

        # ==================================================
        # NORMALIZE COLUMN NAMES
        # ==================================================

        normalized_columns = {
            str(col).lower().strip(): col
            for col in original_columns
            if col
        }

        def find_column(possible_names):

            for name in possible_names:

                if name in normalized_columns:
                    return normalized_columns[name]

            return None

        # ==================================================
        # DETECT COLUMNS
        # ==================================================

        date_column = find_column([
            "date",
            "transaction date",
            "txn date",
            "value date",
            "transaction_date"
        ])

        description_column = find_column([
            "description",
            "narration",
            "particulars",
            "merchant",
            "remarks",
            "details",
            "transaction details"
        ])

        amount_column = find_column([
            "amount",
            "transaction amount",
            "transaction_amount"
        ])

        debit_column = find_column([
            "debit",
            "withdrawal",
            "withdrawal amount",
            "debit amount",
            "debit_amount"
        ])

        credit_column = find_column([
            "credit",
            "deposit",
            "deposit amount",
            "credit amount",
            "credit_amount"
        ])

        type_column = find_column([
            "type",
            "dr/cr",
            "dr / cr",
            "transaction type",
            "transaction_type"
        ])

        # ==================================================
        # REQUIRED COLUMNS
        # ==================================================

        if not date_column:

            flash(
                "Could not find a date column.",
                "error"
            )

            return redirect(
                url_for("upload_statement_page")
            )

        if not description_column:

            flash(
                "Could not find a description column.",
                "error"
            )

            return redirect(
                url_for("upload_statement_page")
            )

        if (
            not amount_column
            and not debit_column
            and not credit_column
        ):

            flash(
                "Could not find amount/debit/credit columns.",
                "error"
            )

            return redirect(
                url_for("upload_statement_page")
            )

        # ==================================================
        # DATE FUNCTION
        # ==================================================

        def normalize_date(date_text):

            date_text = str(
                date_text or ""
            ).strip()

            if not date_text:
                return ""

            if " " in date_text:
                date_text = date_text.split(" ")[0]

            possible_formats = [
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d/%m/%Y",
                "%Y/%m/%d",
                "%d-%m-%y",
                "%d/%m/%y"
            ]

            for date_format in possible_formats:

                try:

                    parsed_date = datetime.strptime(
                        date_text,
                        date_format
                    )

                    return parsed_date.strftime(
                        "%Y-%m-%d"
                    )

                except ValueError:
                    continue

            return date_text

        # ==================================================
        # AMOUNT FUNCTION
        # ==================================================

        def clean_amount(amount_text):

            amount_text = str(
                amount_text or ""
            )

            return (
                amount_text
                .replace("₹", "")
                .replace(",", "")
                .replace("INR", "")
                .replace("Rs.", "")
                .replace("Rs", "")
                .strip()
            )

        # ==================================================
        # BUILD TRANSACTIONS
        # ==================================================

        transactions = []

        for row in rows:

            transaction = {
                "date": "",
                "description": "",
                "amount": "",
                "type": "",
                "category": "Other"
            }

            raw_date = str(
                row.get(date_column, "") or ""
            ).strip()

            transaction["date"] = normalize_date(
                raw_date
            )

            transaction["description"] = str(
                row.get(
                    description_column,
                    ""
                ) or ""
            ).strip()

            # ==================================================
            # AMOUNT + TYPE
            # ==================================================

            if amount_column:

                transaction["amount"] = clean_amount(
                    row.get(
                        amount_column,
                        ""
                    )
                )

                if type_column:

                    transaction["type"] = str(
                        row.get(
                            type_column,
                            ""
                        ) or ""
                    ).strip()

            # ==================================================
            # DEBIT / CREDIT COLUMNS
            # ==================================================

            else:

                debit_value = ""
                credit_value = ""

                if debit_column:

                    debit_value = clean_amount(
                        row.get(
                            debit_column,
                            ""
                        )
                    )

                if credit_column:

                    credit_value = clean_amount(
                        row.get(
                            credit_column,
                            ""
                        )
                    )

                if debit_value:

                    transaction["amount"] = debit_value
                    transaction["type"] = "Debit"

                elif credit_value:

                    transaction["amount"] = credit_value
                    transaction["type"] = "Credit"

            # ==================================================
            # NORMALIZE TYPE
            # ==================================================

            transaction_type = (
                transaction["type"]
                .strip()
                .lower()
            )

            if transaction_type in [
                "credit",
                "cr",
                "c",
                "income"
            ]:

                transaction["type"] = "Credit"

            elif transaction_type in [
                "debit",
                "dr",
                "d",
                "expense"
            ]:

                transaction["type"] = "Debit"

            # Skip blank rows

            if (
                not transaction["description"]
                and not transaction["amount"]
            ):
                continue

                       # ==================================================
            # AI CATEGORY
            # ==================================================

            saved_category = get_saved_merchant_category(
                session["user_id"],
                transaction["description"]
            )

            if saved_category:

                transaction["category"] = saved_category

            else:

                transaction["category"] = categorize_transaction(
                    transaction["description"],
                    transaction["type"]
                )

            print(
                transaction["description"],
                transaction["type"],
                "=>",
                transaction["category"]
            )

            transactions.append(
                transaction
            )

        # ==================================================
        # FINAL CHECK
        # ==================================================

        if not transactions:

            flash(
                "No valid transactions found.",
                "error"
            )

            return redirect(
                url_for("upload_statement_page")
            )

        return render_template(
            "statement-preview.html",
            transactions=transactions,
            columns=[
                "date",
                "description",
                "amount",
                "type",
                "category"
            ]
        )

    except Exception as error:

        print(
            "Statement Parse Error:",
            error
        )

        flash(
            "Unable to understand this bank statement.",
            "error"
        )

        return redirect(
            url_for("upload_statement_page")
        )


# ==================================================
# IMPORT STATEMENT TRANSACTIONS
# ==================================================

@app.route("/import-statement-transactions", methods=["POST"])
@login_required
def import_statement_transactions():

    user_id = session["user_id"]

    try:

        transaction_count = int(
            request.form.get("transaction_count", 0)
        )

        if transaction_count <= 0:

            flash(
                "No transactions available to import.",
                "error"
            )

            return redirect(
                url_for("upload_statement_page")
            )

        cursor = db.cursor()

        imported_count = 0
        duplicate_count = 0

        for index in range(transaction_count):

            transaction_date = request.form.get(
                f"date_{index}",
                ""
            ).strip()

            description = request.form.get(
                f"description_{index}",
                ""
            ).strip()

            amount_text = request.form.get(
                f"amount_{index}",
                ""
            ).strip()

            transaction_type = request.form.get(
                f"type_{index}",
                ""
            ).strip().lower()

            category = request.form.get(
                f"category_{index}",
                "Other"
            ).strip()
            save_merchant_category(
    user_id,
    description,
    category
)

            # Clean amount
            amount_text = (
                amount_text
                .replace("₹", "")
                .replace(",", "")
                .strip()
            )

            try:

                amount = float(amount_text)

            except ValueError:

                print(
                    "Skipped invalid amount:",
                    amount_text
                )

                continue

            if amount <= 0:
                continue


            # ======================================
            # CREDIT → INCOME TABLE
            # ======================================

            if transaction_type in [
                "credit",
                "cr",
                "income"
            ]:

                source = (
                    description
                    if description
                    else category
                )

                # Check duplicate income
                cursor.execute("""
                    SELECT id
                    FROM income
                    WHERE user_id = %s
                      AND source = %s
                      AND amount = %s
                      AND income_date = %s
                    LIMIT 1
                """, (
                    user_id,
                    source,
                    amount,
                    transaction_date
                ))

                existing_income = cursor.fetchone()

                if existing_income:

                    duplicate_count += 1

                    print(
                        "Duplicate income skipped:",
                        source,
                        amount,
                        transaction_date
                    )

                    continue

                cursor.execute("""
                    INSERT INTO income
                    (
                        user_id,
                        source,
                        amount,
                        income_date
                    )
                    VALUES (%s, %s, %s, %s)
                """, (
                    user_id,
                    source,
                    amount,
                    transaction_date
                ))

                imported_count += 1


            # ======================================
            # DEBIT → EXPENSE TABLE
            # ======================================

            elif transaction_type in [
                "debit",
                "dr",
                "expense"
            ]:

                # Check duplicate expense
                cursor.execute("""
                    SELECT id
                    FROM expenses
                    WHERE user_id = %s
                      AND category = %s
                      AND amount = %s
                      AND description = %s
                      AND expense_date = %s
                    LIMIT 1
                """, (
                    user_id,
                    category,
                    amount,
                    description,
                    transaction_date
                ))

                existing_expense = cursor.fetchone()

                if existing_expense:

                    duplicate_count += 1

                    print(
                        "Duplicate expense skipped:",
                        description,
                        amount,
                        transaction_date
                    )

                    continue

                cursor.execute("""
                    INSERT INTO expenses
                    (
                        user_id,
                        category,
                        amount,
                        description,
                        expense_date
                    )
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    user_id,
                    category,
                    amount,
                    description,
                    transaction_date
                ))

                imported_count += 1


        db.commit()
        cursor.close()

        flash(
            f"{imported_count} new transactions imported. "
            f"{duplicate_count} duplicates skipped.",
            "success"
        )

        return redirect(
            url_for("dashboard")
        )


    except Exception as error:

        db.rollback()

        print(
            "Statement Import Error:",
            error
        )

        flash(
            "Unable to import transactions. Please try again.",
            "error"
        )

        return redirect(
            url_for("upload_statement_page")
        )
# ==================================================
# REPORTS
# ==================================================

@app.route("/reports")
@login_required
def reports():

    user_id = session["user_id"]

    # ==========================================
    # SELECT MONTH + YEAR
    # ==========================================

    selected_month = request.args.get(
        "month",
        datetime.now().strftime("%m")
    )

    selected_year = request.args.get(
        "year",
        str(datetime.now().year)
    )

    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)

    except ValueError:
        selected_month = datetime.now().month
        selected_year = datetime.now().year


    cursor = db.cursor(dictionary=True)

    # ==========================================
    # CATEGORY TOTALS FOR SELECTED MONTH
    # ==========================================

    cursor.execute("""
        SELECT
    category,
    SUM(amount) AS total,
    COUNT(*) AS transaction_count
FROM expenses
WHERE user_id = %s
  AND MONTH(expense_date) = %s
  AND YEAR(expense_date) = %s
GROUP BY category
ORDER BY total DESC
    """, (
        user_id,
        selected_month,
        selected_year
    ))

    report_data = cursor.fetchall()


    # ==========================================
    # MONTH TOTAL
    # ==========================================

    cursor.execute("""
        SELECT
            COALESCE(SUM(amount), 0) AS total
        FROM expenses
        WHERE user_id = %s
          AND MONTH(expense_date) = %s
          AND YEAR(expense_date) = %s
    """, (
        user_id,
        selected_month,
        selected_year
    ))

    total_report_expense = float(
        cursor.fetchone()["total"]
    )


    # ==========================================
    # TRANSACTION COUNT
    # ==========================================

    cursor.execute("""
        SELECT
            COUNT(*) AS total
        FROM expenses
        WHERE user_id = %s
          AND MONTH(expense_date) = %s
          AND YEAR(expense_date) = %s
    """, (
        user_id,
        selected_month,
        selected_year
    ))

    monthly_transaction_count = (
        cursor.fetchone()["total"]
    )

    cursor.close()


    # ==========================================
    # CHART DATA
    # ==========================================

    labels = []
    values = []

    for row in report_data:

        labels.append(
            row["category"].title()
        )

        values.append(
            float(row["total"])
        )


    # ==========================================
    # CATEGORY SUMMARY
    # ==========================================

    category_summary = []

    for row in report_data:

        amount = float(
            row["total"]
        )

        if total_report_expense > 0:

            percentage = (
                amount
                / total_report_expense
            ) * 100

        else:

            percentage = 0

        category_summary.append({
    "category": row["category"].title(),
    "amount": amount,
    "percentage": percentage,
    "transaction_count": row["transaction_count"]
})


    # ==========================================
    # TOP CATEGORY
    # ==========================================

    if report_data:

        top_category = (
            report_data[0]["category"].title()
        )

        top_category_amount = float(
            report_data[0]["total"]
        )

    else:

        top_category = "No Data"
        top_category_amount = 0


    # ==========================================
    # MONTH NAME
    # ==========================================

    selected_month_name = datetime(
        selected_year,
        selected_month,
        1
    ).strftime("%B")


    return render_template(
        "reports.html",

        labels=labels,
        values=values,

        total_report_expense=total_report_expense,

        top_category=top_category,
        top_category_amount=top_category_amount,

        category_summary=category_summary,

        monthly_transaction_count=monthly_transaction_count,

        selected_month=selected_month,
        selected_year=selected_year,
        selected_month_name=selected_month_name
    )
# ==================================================
# DOWNLOAD PDF REPORT
# ==================================================

@app.route("/download-report")
@login_required
def download_report():

    user_id = session["user_id"]

    # ==========================================
    # GET MONTH + YEAR
    # ==========================================

    selected_month = request.args.get(
        "month",
        datetime.now().strftime("%m")
    )

    selected_year = request.args.get(
        "year",
        str(datetime.now().year)
    )

    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)

    except ValueError:
        selected_month = datetime.now().month
        selected_year = datetime.now().year

    selected_month_name = datetime(
        selected_year,
        selected_month,
        1
    ).strftime("%B")


    # ==========================================
    # GET MONTHLY CATEGORY TOTALS
    # ==========================================

    cursor = db.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            category,
            SUM(amount) AS total
        FROM expenses
        WHERE user_id = %s
          AND MONTH(expense_date) = %s
          AND YEAR(expense_date) = %s
        GROUP BY category
        ORDER BY total DESC
    """, (
        user_id,
        selected_month,
        selected_year
    ))

    report_data = cursor.fetchall()


    # ==========================================
    # MONTHLY TRANSACTION COUNT
    # ==========================================

    cursor.execute("""
        SELECT
            COUNT(*) AS total
        FROM expenses
        WHERE user_id = %s
          AND MONTH(expense_date) = %s
          AND YEAR(expense_date) = %s
    """, (
        user_id,
        selected_month,
        selected_year
    ))

    transaction_count = cursor.fetchone()["total"]

    cursor.close()


    # ==========================================
    # CREATE PDF
    # ==========================================

    pdf_buffer = BytesIO()

    pdf = SimpleDocTemplate(
        pdf_buffer,
        pagesize=(8.27 * inch, 11.69 * inch),
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    elements = []


    # ==========================================
    # TITLE
    # ==========================================

    elements.append(
        Paragraph(
            "Smart Expense Tracker AI",
            styles["Title"]
        )
    )

    elements.append(
        Spacer(
            1,
            8
        )
    )

    elements.append(
        Paragraph(
            f"{selected_month_name} {selected_year} Expense Report",
            styles["Heading2"]
        )
    )

    elements.append(
        Spacer(
            1,
            8
        )
    )


    # ==========================================
    # USER DETAILS
    # ==========================================

    elements.append(
        Paragraph(
            f"User: {session.get('fullname', 'User')}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"Expense Transactions: {transaction_count}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            "Generated on: "
            + datetime.now().strftime(
                "%d %b %Y, %I:%M %p"
            ),
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(
            1,
            20
        )
    )


    # ==========================================
    # TABLE
    # ==========================================

    table_data = [
        [
            "Category",
            "Amount",
            "Percentage"
        ]
    ]

    total_expense = sum(
        float(row["total"])
        for row in report_data
    )


    for row in report_data:

        amount = float(
            row["total"]
        )

        if total_expense > 0:

            percentage = (
                amount
                / total_expense
            ) * 100

        else:

            percentage = 0

        table_data.append([
            row["category"].title(),
            f"Rs. {amount:,.0f}",
            f"{percentage:.1f}%"
        ])


    table_data.append([
        "Total Expense",
        f"Rs. {total_expense:,.0f}",
        "100.0%"
        if total_expense > 0
        else "0.0%"
    ])


    report_table = Table(
        table_data,
        colWidths=[
            2.7 * inch,
            2.0 * inch,
            1.5 * inch
        ]
    )


    report_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1f4f99")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.grey
            ),

            (
                "BACKGROUND",
                (0, 1),
                (-1, -2),
                colors.whitesmoke
            ),

            (
                "BACKGROUND",
                (0, -1),
                (-1, -1),
                colors.lightgrey
            ),

            (
                "FONTNAME",
                (0, -1),
                (-1, -1),
                "Helvetica-Bold"
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                10
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                10
            )
        ])
    )

    elements.append(
        report_table
    )


    # ==========================================
    # BUILD PDF
    # ==========================================

    pdf.build(
        elements
    )

    pdf_buffer.seek(0)


    # ==========================================
    # DOWNLOAD
    # ==========================================

    filename = (
        f"expense_report_"
        f"{selected_month_name.lower()}_"
        f"{selected_year}.pdf"
    )

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )
# ==================================================
# BUDGET PAGE
# ==================================================

@app.route("/budget")
@login_required
def budget():

    user_id = session["user_id"]

    # Get selected month/year from URL
    selected_month = request.args.get(
        "month",
        datetime.now().month
    )

    selected_year = request.args.get(
        "year",
        datetime.now().year
    )

    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)

    except (ValueError, TypeError):
        selected_month = datetime.now().month
        selected_year = datetime.now().year

    # Validate month
    if selected_month < 1 or selected_month > 12:
        selected_month = datetime.now().month

    # Convert 5 -> May, 8 -> August, etc.
    selected_month_name = datetime(
        selected_year,
        selected_month,
        1
    ).strftime("%B")

    cursor = db.cursor(dictionary=True)

    try:

        # ==================================================
        # GET SAVED BUDGETS
        # budgets.month stores names like "May", "August"
        # ==================================================

        cursor.execute("""
            SELECT
                category,
                budget_amount
            FROM budgets
            WHERE user_id = %s
              AND month = %s
              AND year = %s
            ORDER BY id DESC
        """, (
            user_id,
            selected_month_name,
            selected_year
        ))

        budget_rows = cursor.fetchall()

        # ==================================================
        # KEEP LATEST BUDGET PER CATEGORY
        # ==================================================

        latest_budgets = {}

        for row in budget_rows:

            category = str(
                row["category"]
            ).strip().title()

            if category not in latest_budgets:

                latest_budgets[category] = float(
                    row["budget_amount"]
                )

        # ==================================================
        # GET EXPENSES FOR SELECTED MONTH
        #
        # IMPORTANT:
        # MONTH(expense_date) needs 1-12,
        # NOT "May", "August", etc.
        # ==================================================

        cursor.execute("""
            SELECT
                category,
                COALESCE(SUM(amount), 0) AS spent
            FROM expenses
            WHERE user_id = %s
              AND MONTH(expense_date) = %s
              AND YEAR(expense_date) = %s
            GROUP BY category
        """, (
            user_id,
            selected_month,
            selected_year
        ))

        expense_rows = cursor.fetchall()

    finally:
        cursor.close()

    # ==================================================
    # CREATE EXPENSE LOOKUP
    # ==================================================

    spent_by_category = {}

    for row in expense_rows:

        category = str(
            row["category"]
        ).strip().title()

        spent_by_category[category] = float(
            row["spent"] or 0
        )

    # ==================================================
    # BUILD BUDGET CARDS
    # ==================================================

    budget_details = []

    for category, budget_amount in latest_budgets.items():

        spent = spent_by_category.get(
            category,
            0
        )

        remaining = budget_amount - spent

        if budget_amount > 0:

            percentage = (
                spent / budget_amount
            ) * 100

        else:

            percentage = 0

        budget_details.append({

            "category": category,

            "budget": budget_amount,

            "spent": spent,

            "remaining": remaining,

            "percentage": percentage,

            "progress_width": min(
                percentage,
                100
            ),

            "exceeded": spent > budget_amount
        })

    # ==================================================
    # TOTALS
    # ==================================================

    total_budget = sum(
        item["budget"]
        for item in budget_details
    )

    total_spent = sum(
        item["spent"]
        for item in budget_details
    )

    total_remaining = (
        total_budget - total_spent
    )

    if total_budget > 0:

        overall_percentage = (
            total_spent / total_budget
        ) * 100

    else:

        overall_percentage = 0

    # ==================================================
    # RENDER
    # ==================================================

    return render_template(
        "budget.html",

        budget_details=budget_details,

        selected_month=selected_month,
        selected_year=selected_year,
        selected_month_name=selected_month_name,

        current_month=selected_month_name,
        current_year=selected_year,

        total_budget=total_budget,
        total_spent=total_spent,
        total_remaining=total_remaining,

        overall_percentage=overall_percentage
    )


# ==================================================
# SAVE BUDGET
# ==================================================

@app.route("/save-budget", methods=["POST"])
@login_required
def save_budget():

    user_id = session["user_id"]

    # ==================================================
    # GET SELECTED MONTH/YEAR
    # ==================================================

    selected_month = request.form.get(
        "selected_month",
        datetime.now().month
    )

    selected_year = request.form.get(
        "selected_year",
        datetime.now().year
    )

    try:

        selected_month = int(
            selected_month
        )

        selected_year = int(
            selected_year
        )

    except (ValueError, TypeError):

        selected_month = datetime.now().month
        selected_year = datetime.now().year

    # Validate month
    if selected_month < 1 or selected_month > 12:
        selected_month = datetime.now().month

    # Convert 5 -> May
    month_name = datetime(
        selected_year,
        selected_month,
        1
    ).strftime("%B")

    # ==================================================
    # GET FORM VALUES
    # ==================================================

    food_budget = request.form.get(
        "food_budget",
        ""
    ).strip()

    petrol_budget = request.form.get(
        "petrol_budget",
        ""
    ).strip()

    studies_budget = request.form.get(
        "studies_budget",
        ""
    ).strip()

    rent_budget = request.form.get(
        "rent_budget",
        ""
    ).strip()

    budget_data = [
        ("Food", food_budget),
        ("Petrol", petrol_budget),
        ("Studies", studies_budget),
        ("Rent", rent_budget)
    ]

    cursor = db.cursor()

    saved_count = 0

    try:

        for category, amount_text in budget_data:

            # Blank fields are allowed
            if not amount_text:
                continue

            try:

                amount = float(
                    amount_text
                )

                if amount < 0:
                    continue

            except (ValueError, TypeError):
                continue

            # ==================================================
            # REMOVE OLD VALUE FOR SAME CATEGORY/MONTH/YEAR
            # ==================================================

            cursor.execute("""
                DELETE FROM budgets
                WHERE user_id = %s
                  AND LOWER(category) = LOWER(%s)
                  AND month = %s
                  AND year = %s
            """, (
                user_id,
                category,
                month_name,
                selected_year
            ))

            # ==================================================
            # INSERT NEW VALUE
            # ==================================================

            cursor.execute("""
                INSERT INTO budgets
                (
                    user_id,
                    category,
                    budget_amount,
                    month,
                    year
                )
                VALUES (%s, %s, %s, %s, %s)
            """, (
                user_id,
                category,
                amount,
                month_name,
                selected_year
            ))

            saved_count += 1

        db.commit()

        if saved_count > 0:

            flash(
                f"Budget saved for "
                f"{month_name} {selected_year}.",
                "success"
            )

        else:

            flash(
                "Please enter at least one budget amount.",
                "error"
            )

    except Exception as error:

        db.rollback()

        print(
            "Save Budget Error:",
            error
        )

        flash(
            "Unable to save budget.",
            "error"
        )

    finally:

        cursor.close()

    # Return to SAME selected month
    return redirect(
        url_for(
            "budget",
            month=selected_month,
            year=selected_year
        )
    )

# ==================================================
# ADD CUSTOM BUDGET
# ==================================================

@app.route("/add-custom-budget", methods=["POST"])
@login_required
def add_custom_budget():

    user_id = session["user_id"]

    category = request.form.get(
        "category",
        ""
    ).strip().title()

    budget_amount = request.form.get(
        "budget_amount",
        ""
    ).strip()


    selected_month = request.form.get(
        "selected_month",
        datetime.now().month
    )

    selected_year = request.form.get(
        "selected_year",
        datetime.now().year
    )


    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)

    except (ValueError, TypeError):
        selected_month = datetime.now().month
        selected_year = datetime.now().year


    month_name = datetime(
        selected_year,
        selected_month,
        1
    ).strftime("%B")


    if not category:

        flash(
            "Category is required.",
            "error"
        )

        return redirect(
            url_for(
                "budget",
                month=selected_month,
                year=selected_year
            )
        )


    try:

        budget_amount = float(
            budget_amount
        )

        if budget_amount < 0:
            raise ValueError

    except ValueError:

        flash(
            "Enter a valid budget amount.",
            "error"
        )

        return redirect(
            url_for(
                "budget",
                month=selected_month,
                year=selected_year
            )
        )


    cursor = db.cursor()

    try:

        cursor.execute("""
            DELETE FROM budgets
            WHERE user_id = %s
              AND LOWER(category) = LOWER(%s)
              AND month = %s
              AND year = %s
        """, (
            user_id,
            category,
            month_name,
            selected_year
        ))


        cursor.execute("""
            INSERT INTO budgets
            (
                user_id,
                category,
                budget_amount,
                month,
                year
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            user_id,
            category,
            budget_amount,
            month_name,
            selected_year
        ))


        db.commit()

        flash(
            f"{category} budget added for "
            f"{month_name} {selected_year}.",
            "success"
        )


    except Exception as error:

        db.rollback()

        print(
            "Custom Budget Error:",
            error
        )

        flash(
            "Unable to add budget.",
            "error"
        )


    finally:
        cursor.close()


    return redirect(
        url_for(
            "budget",
            month=selected_month,
            year=selected_year
        )
    )


# ==================================================
# DELETE BUDGET
# ==================================================

@app.route(
    "/delete-budget/<category>",
    methods=["POST"]
)
@login_required
def delete_budget(category):

    user_id = session["user_id"]

    selected_month = request.form.get(
        "selected_month",
        datetime.now().month
    )

    selected_year = request.form.get(
        "selected_year",
        datetime.now().year
    )


    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)

    except (ValueError, TypeError):
        selected_month = datetime.now().month
        selected_year = datetime.now().year


    month_name = datetime(
        selected_year,
        selected_month,
        1
    ).strftime("%B")


    cursor = db.cursor()

    try:

        cursor.execute("""
            DELETE FROM budgets
            WHERE user_id = %s
              AND category = %s
              AND month = %s
              AND year = %s
        """, (
            user_id,
            category,
            month_name,
            selected_year
        ))

        db.commit()

        flash(
            f"{category} budget deleted.",
            "success"
        )


    except Exception as error:

        db.rollback()

        print(
            "Delete Budget Error:",
            error
        )

        flash(
            "Unable to delete budget.",
            "error"
        )


    finally:
        cursor.close()


    return redirect(
        url_for(
            "budget",
            month=selected_month,
            year=selected_year
        )
    )

# ==================================================
# EDIT INCOME PAGE
# ==================================================

@app.route("/edit-income/<int:income_id>")
@login_required
def edit_income_page(income_id):

    user_id = session["user_id"]

    cursor = db.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            id,
            source,
            amount,
            income_date
        FROM income
        WHERE id = %s
          AND user_id = %s
    """, (
        income_id,
        user_id
    ))

    income_item = cursor.fetchone()

    cursor.close()

    if not income_item:
        return redirect(url_for("dashboard"))

    return render_template(
        "edit-income.html",
        income_item=income_item
    )


# ==================================================
# UPDATE INCOME
# ==================================================

@app.route("/update-income/<int:income_id>", methods=["POST"])
@login_required
def update_income(income_id):

    user_id = session["user_id"]

    source = request.form["source"]
    amount = request.form["amount"]
    income_date = request.form["income_date"]

    cursor = db.cursor()

    cursor.execute("""
        UPDATE income
        SET
            source = %s,
            amount = %s,
            income_date = %s
        WHERE id = %s
          AND user_id = %s
    """, (
        source,
        amount,
        income_date,
        income_id,
        user_id
    ))

    db.commit()
    cursor.close()
    flash("Income updated successfully!", "success")

    return redirect(url_for("dashboard"))

# ==================================================
# AI ASSISTANT
# ==================================================

@app.route("/ai", methods=["GET", "POST"])
@login_required
def ai():

    user_id = session["user_id"]

    # ==================================================
    # SELECTED MONTH / YEAR
    # ==================================================

    selected_month = request.values.get(
        "month",
        datetime.now().month
    )

    selected_year = request.values.get(
        "year",
        datetime.now().year
    )

    try:

        selected_month = int(selected_month)
        selected_year = int(selected_year)

    except (ValueError, TypeError):

        selected_month = datetime.now().month
        selected_year = datetime.now().year

    if selected_month < 1 or selected_month > 12:
        selected_month = datetime.now().month

    selected_month_name = datetime(
        selected_year,
        selected_month,
        1
    ).strftime("%B")


    cursor = db.cursor(dictionary=True)

    try:

        # ==================================================
        # MONTHLY INCOME
        # ==================================================

        cursor.execute("""
            SELECT
                COALESCE(SUM(amount), 0) AS total
            FROM income
            WHERE user_id = %s
              AND MONTH(income_date) = %s
              AND YEAR(income_date) = %s
        """, (
            user_id,
            selected_month,
            selected_year
        ))

        total_income = float(
            cursor.fetchone()["total"] or 0
        )


        # ==================================================
        # MONTHLY EXPENSE
        # ==================================================

        cursor.execute("""
            SELECT
                COALESCE(SUM(amount), 0) AS total
            FROM expenses
            WHERE user_id = %s
              AND MONTH(expense_date) = %s
              AND YEAR(expense_date) = %s
        """, (
            user_id,
            selected_month,
            selected_year
        ))

        total_expense = float(
            cursor.fetchone()["total"] or 0
        )


        # ==================================================
        # MONTHLY CATEGORY TOTALS
        # ==================================================

        cursor.execute("""
            SELECT
                category,
                SUM(amount) AS total
            FROM expenses
            WHERE user_id = %s
              AND MONTH(expense_date) = %s
              AND YEAR(expense_date) = %s
            GROUP BY category
            ORDER BY total DESC
        """, (
            user_id,
            selected_month,
            selected_year
        ))

        category_data = cursor.fetchall()


        # ==================================================
        # BUDGETS FOR SELECTED MONTH
        # ==================================================

        cursor.execute("""
            SELECT
                category,
                budget_amount
            FROM budgets
            WHERE user_id = %s
              AND month = %s
              AND year = %s
            ORDER BY id DESC
        """, (
            user_id,
            selected_month_name,
            selected_year
        ))

        budget_rows = cursor.fetchall()

    finally:

        cursor.close()


    # ==================================================
    # CATEGORY SPENDING DICTIONARY
    # ==================================================

    monthly_spending = {}

    for row in category_data:

        category_name = str(
            row["category"]
        ).strip().lower()

        monthly_spending[category_name] = float(
            row["total"] or 0
        )


    # ==================================================
    # BUDGET DICTIONARY
    # ==================================================

    budget_by_category = {}

    for row in budget_rows:

        category_name = str(
            row["category"]
        ).strip().lower()

        # Keep latest budget if duplicate records exist
        if category_name not in budget_by_category:

            budget_by_category[category_name] = float(
                row["budget_amount"] or 0
            )


    # ==================================================
    # FINANCIAL CALCULATIONS
    # ==================================================

    savings = total_income - total_expense

    total_budget = sum(
        budget_by_category.values()
    )

    remaining_budget = (
        total_budget - total_expense
    )

    exceeded_categories = []

    for category_name, budget_amount in budget_by_category.items():

        spent_amount = monthly_spending.get(
            category_name,
            0
        )

        if spent_amount > budget_amount:

            exceeded_categories.append(
                category_name.title()
            )


    # ==================================================
    # DEFAULT AI SUGGESTION
    # ==================================================

    question = ""

    if category_data:

        highest = category_data[0]

        highest_category = str(
            highest["category"]
        ).title()

        highest_total = float(
            highest["total"] or 0
        )

        possible_saving = (
            highest_total * 0.15
        )

        suggestion = (
            f"In {selected_month_name} {selected_year}, "
            f"you spent ₹{total_expense:,.0f}. "
            f"Your highest spending category was "
            f"{highest_category} at ₹{highest_total:,.0f}. "
            f"Reducing {highest_category} spending by 15% "
            f"could save approximately "
            f"₹{possible_saving:,.0f}."
        )

    else:

        suggestion = (
            f"No expense data was found for "
            f"{selected_month_name} {selected_year}. "
            f"Upload a statement or add expenses to receive "
            f"financial insights."
        )


    # ==================================================
    # ANSWER USER QUESTION
    # ==================================================

    if request.method == "POST":

        question = request.form.get(
            "question",
            ""
        ).strip()

        question_lower = question.lower()


        if not question:

            suggestion = (
                "Please enter a financial question."
            )


        # ==================================================
        # CATEGORY-SPECIFIC QUESTION
        #
        # Example:
        # How much did I spend on petrol?
        # How much did I spend on food?
        # ==================================================

        else:

            matched_category = None

            for category_name in monthly_spending:

                if category_name in question_lower:

                    matched_category = category_name
                    break


            if (
                matched_category
                and any(
                    word in question_lower
                    for word in [
                        "spend",
                        "spent",
                        "expense",
                        "expenses",
                        "how much"
                    ]
                )
            ):

                category_amount = monthly_spending[
                    matched_category
                ]

                percentage = 0

                if total_expense > 0:

                    percentage = (
                        category_amount /
                        total_expense
                    ) * 100

                suggestion = (
                    f"In {selected_month_name} {selected_year}, "
                    f"you spent ₹{category_amount:,.0f} on "
                    f"{matched_category.title()}. "
                    f"That represents approximately "
                    f"{percentage:.1f}% of your total "
                    f"monthly expenses."
                )


            # ==================================================
            # INCOME
            # ==================================================

            elif "income" in question_lower:

                suggestion = (
                    f"Your income for "
                    f"{selected_month_name} {selected_year} "
                    f"is ₹{total_income:,.0f}."
                )


            # ==================================================
            # HIGHEST CATEGORY
            # ==================================================

            elif (
                "highest" in question_lower
                or "top category" in question_lower
                or "most spent" in question_lower
                or "spend most" in question_lower
                or "spent most" in question_lower
            ):

                if category_data:

                    highest = category_data[0]

                    highest_category = str(
                        highest["category"]
                    ).title()

                    highest_amount = float(
                        highest["total"] or 0
                    )

                    suggestion = (
                        f"Your highest spending category in "
                        f"{selected_month_name} {selected_year} "
                        f"is {highest_category} at "
                        f"₹{highest_amount:,.0f}."
                    )

                else:

                    suggestion = (
                        f"No expenses were found for "
                        f"{selected_month_name} {selected_year}."
                    )


            # ==================================================
            # SAVINGS
            # ==================================================

            elif (
                "saving" in question_lower
                or "savings" in question_lower
            ):

                if savings >= 0:

                    suggestion = (
                        f"For {selected_month_name} {selected_year}, "
                        f"your income is ₹{total_income:,.0f} "
                        f"and your expenses are ₹{total_expense:,.0f}. "
                        f"Your remaining savings are "
                        f"₹{savings:,.0f}."
                    )

                else:

                    suggestion = (
                        f"For {selected_month_name} {selected_year}, "
                        f"your expenses exceed your recorded income "
                        f"by ₹{abs(savings):,.0f}."
                    )


            # ==================================================
            # BUDGET
            # ==================================================

            elif (
                "budget" in question_lower
                or "exceeded" in question_lower
                or "over budget" in question_lower
            ):

                matched_budget_category = None

                for category_name in budget_by_category:

                    if category_name in question_lower:

                        matched_budget_category = (
                            category_name
                        )

                        break


                # Specific category budget
                if matched_budget_category:

                    category_budget = (
                        budget_by_category[
                            matched_budget_category
                        ]
                    )

                    category_spent = (
                        monthly_spending.get(
                            matched_budget_category,
                            0
                        )
                    )

                    category_remaining = (
                        category_budget -
                        category_spent
                    )

                    if category_remaining >= 0:

                        suggestion = (
                            f"Your "
                            f"{matched_budget_category.title()} "
                            f"budget for "
                            f"{selected_month_name} {selected_year} "
                            f"is ₹{category_budget:,.0f}. "
                            f"You have spent "
                            f"₹{category_spent:,.0f}, leaving "
                            f"₹{category_remaining:,.0f}."
                        )

                    else:

                        suggestion = (
                            f"Your "
                            f"{matched_budget_category.title()} "
                            f"budget is ₹{category_budget:,.0f}, "
                            f"but you have spent "
                            f"₹{category_spent:,.0f}. "
                            f"You exceeded the budget by "
                            f"₹{abs(category_remaining):,.0f}."
                        )


                elif total_budget == 0:

                    suggestion = (
                        f"You have not saved any budgets for "
                        f"{selected_month_name} {selected_year}."
                    )


                else:

                    suggestion = (
                        f"Your total saved budget for "
                        f"{selected_month_name} {selected_year} "
                        f"is ₹{total_budget:,.0f}."
                    )

                    if exceeded_categories:

                        suggestion += (
                            " Exceeded categories: "
                            + ", ".join(
                                exceeded_categories
                            )
                            + "."
                        )

                    else:

                        suggestion += (
                            " All saved category budgets "
                            "are currently on track."
                        )


            # ==================================================
            # REDUCE / RECOMMENDATION
            # ==================================================

            elif any(
                phrase in question_lower
                for phrase in [
                    "reduce",
                    "recommend",
                    "recommendation",
                    "suggest",
                    "control spending",
                    "save money"
                ]
            ):

                if category_data:

                    highest = category_data[0]

                    highest_category = str(
                        highest["category"]
                    ).title()

                    highest_amount = float(
                        highest["total"] or 0
                    )

                    reduction = (
                        highest_amount * 0.15
                    )

                    suggestion = (
                        f"Your largest expense in "
                        f"{selected_month_name} {selected_year} "
                        f"is {highest_category} at "
                        f"₹{highest_amount:,.0f}. "
                        f"A 15% reduction in this category "
                        f"would save approximately "
                        f"₹{reduction:,.0f}."
                    )

                else:

                    suggestion = (
                        "I need expense data before I can "
                        "recommend where to reduce spending."
                    )


            # ==================================================
            # TOTAL EXPENSE
            # ==================================================

            elif any(
                word in question_lower
                for word in [
                    "expense",
                    "expenses",
                    "spend",
                    "spent",
                    "spending"
                ]
            ):

                suggestion = (
                    f"Your total expenses for "
                    f"{selected_month_name} {selected_year} "
                    f"are ₹{total_expense:,.0f}."
                )


            # ==================================================
            # GENERAL SUMMARY
            # ==================================================

            elif (
                "summary" in question_lower
                or "analyse" in question_lower
                or "analyze" in question_lower
                or "analysis" in question_lower
            ):

                if category_data:

                    highest = category_data[0]

                    highest_category = str(
                        highest["category"]
                    ).title()

                    highest_amount = float(
                        highest["total"] or 0
                    )

                    suggestion = (
                        f"{selected_month_name} {selected_year} summary:\n"
                        f"Income: ₹{total_income:,.0f}\n"
                        f"Expenses: ₹{total_expense:,.0f}\n"
                        f"Savings: ₹{savings:,.0f}\n"
                        f"Highest category: "
                        f"{highest_category} "
                        f"(₹{highest_amount:,.0f})"
                    )

                else:

                    suggestion = (
                        f"No financial activity is available for "
                        f"{selected_month_name} {selected_year}."
                    )


            # ==================================================
            # UNKNOWN QUESTION
            # ==================================================

            else:

                suggestion = (
                    "I can analyse your monthly income, expenses, "
                    "savings, category spending and budgets. "
                    "Try asking: 'How much did I spend on petrol?', "
                    "'What is my highest expense?', "
                    "'Did I exceed my budget?' or "
                    "'Give me a monthly summary.'"
                )


    # ==================================================
    # RENDER
    # ==================================================

    return render_template(
        "ai-assistant.html",

        suggestion=suggestion,
        question=question,

        selected_month=selected_month,
        selected_year=selected_year,
        selected_month_name=selected_month_name,

        total_income=total_income,
        total_expense=total_expense,
        savings=savings
    )
# ==================================================
# SIGNUP
# ==================================================

@app.route("/signup", methods=["GET", "POST"])
def signup():

    if request.method == "POST":

        fullname = request.form["fullname"]
        email = request.form["email"]
        phone = request.form["phone"]
        password = request.form["password"]
        hashed_password = generate_password_hash(password)
        confirm_password = request.form["confirm_password"]

        if password != confirm_password:

            return render_template(
                "signup.html",
                error="Passwords do not match."
            )

        cursor = db.cursor()

        try:

            cursor.execute("""
                INSERT INTO users
                (
                    fullname,
                    email,
                    phone,
                    password
                )
                VALUES (%s, %s, %s, %s)
            """, (
                fullname,
                email,
                phone,
                hashed_password
            ))

            db.commit()

        except mysql.connector.IntegrityError:

            cursor.close()

            return render_template(
                "signup.html",
                error="Email address already exists."
            )

        cursor.close()

        return redirect(url_for("login"))

    return render_template("signup.html")

# ==================================================
# ADD INCOME PAGE
# ==================================================

@app.route("/add-income")
@login_required
def add_income_page():

    return render_template("add-income.html")

# ==================================================
# SAVE INCOME
# ==================================================

@app.route("/add_income", methods=["POST"])
@login_required
def add_income():

    user_id = session["user_id"]

    source = request.form["source"].strip()
    amount = request.form["amount"]
    income_date = request.form["income_date"]

    # Validate source
    if not source:

        flash(
            "Income source is required.",
            "error"
        )

        return redirect(
            url_for("add_income_page")
        )

    # Validate amount
    try:

        amount = float(amount)

        if amount <= 0:
            raise ValueError

    except ValueError:

        flash(
            "Please enter a valid income amount.",
            "error"
        )

        return redirect(
            url_for("add_income_page")
        )

    # Validate date
    if not income_date:

        flash(
            "Income date is required.",
            "error"
        )

        return redirect(
            url_for("add_income_page")
        )

    cursor = db.cursor()

    cursor.execute("""
        INSERT INTO income
        (
            user_id,
            source,
            amount,
            income_date
        )
        VALUES (%s, %s, %s, %s)
    """, (
        user_id,
        source,
        amount,
        income_date
    ))

    db.commit()
    cursor.close()

    flash(
        "Income added successfully!",
        "success"
    )

    return redirect(
        url_for("dashboard")
    )


# ==================================================
# ADD EXPENSE PAGE
# ==================================================

@app.route("/add-expense")
@login_required
def add_expense_page():

    return render_template("add-expense.html")


# ==================================================
# SAVE EXPENSE
# ==================================================
@app.route("/add_expense", methods=["POST"])
@login_required
def add_expense():

    user_id = session["user_id"]

    category = request.form.get("category", "").strip()
    amount = request.form.get("amount", "").strip()
    description = request.form.get("description", "").strip()
    expense_date = request.form.get("expense_date", "").strip()

    # Validate category
    if not category:
        flash("Expense category is required.", "error")
        return redirect(url_for("add_expense_page"))

    # Validate amount
    try:
        amount = float(amount)

        if amount <= 0:
            raise ValueError

    except ValueError:
        flash("Please enter a valid expense amount.", "error")
        return redirect(url_for("add_expense_page"))

    # Validate date
    if not expense_date:
        flash("Expense date is required.", "error")
        return redirect(url_for("add_expense_page"))

    try:

        cursor = db.cursor()

        cursor.execute("""
            INSERT INTO expenses
            (
                user_id,
                category,
                amount,
                description,
                expense_date
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            user_id,
            category,
            amount,
            description,
            expense_date
        ))

        db.commit()
        cursor.close()

        flash(
            "Expense added successfully!",
            "success"
        )

        return redirect(url_for("dashboard"))

    except Exception as error:

        print("Expense Error:", error)

        flash(
            "Unable to save expense. Please try again.",
            "error"
        )

        return redirect(url_for("add_expense_page"))
    # ==================================================
# DELETE EXPENSE
# ==================================================

@app.route("/delete-expense/<int:expense_id>", methods=["POST"])
@login_required
def delete_expense(expense_id):

    user_id = session["user_id"]

    cursor = db.cursor()

    cursor.execute("""
        DELETE FROM expenses
        WHERE id = %s
        AND user_id = %s
    """, (
        expense_id,
        user_id
    ))

    db.commit()
    cursor.close()
    flash("Expense deleted successfully!", "success")

    return redirect(url_for("dashboard"))

# ==================================================
# EDIT EXPENSE PAGE
# ==================================================

@app.route("/edit-expense/<int:expense_id>")
@login_required
def edit_expense_page(expense_id):

    user_id = session["user_id"]

    cursor = db.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            id,
            category,
            amount,
            description,
            expense_date
        FROM expenses
        WHERE id = %s
          AND user_id = %s
    """, (
        expense_id,
        user_id
    ))

    expense = cursor.fetchone()

    cursor.close()

    if not expense:
        return redirect(url_for("dashboard"))

    return render_template(
        "edit-expense.html",
        expense=expense
    )


# ==================================================
# UPDATE EXPENSE
# ==================================================

@app.route("/update-expense/<int:expense_id>", methods=["POST"])
@login_required
def update_expense(expense_id):

    user_id = session["user_id"]

    category = request.form["category"]
    amount = request.form["amount"]
    description = request.form["description"]
    expense_date = request.form["expense_date"]

    cursor = db.cursor()

    cursor.execute("""
        UPDATE expenses
        SET
            category = %s,
            amount = %s,
            description = %s,
            expense_date = %s
        WHERE id = %s
          AND user_id = %s
    """, (
        category,
        amount,
        description,
        expense_date,
        expense_id,
        user_id
    ))

    db.commit()
    cursor.close()
    flash("Expense updated successfully!", "success")

    return redirect(url_for("dashboard"))

# ==================================================
# LOGOUT
# ==================================================

@app.route("/logout")
def logout():

    session.clear()

    return redirect(url_for("login"))


# ==================================================
# RUN APPLICATION
# ==================================================

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=False
    )